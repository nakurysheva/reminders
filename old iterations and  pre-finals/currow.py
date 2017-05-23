import smtplib
import openpyxl
import sys
# We use email and MIME handling packages to create a nice looking email with subject line
from email.mime.text import MIMEText
from email.header import Header
subject = u'Дедлайн по проекту' 
body = u'Это тестовое письмо оптправлено с помощью smtplib'
msg = MIMEText(body, 'plain', 'utf-8')
msg['Subject'] = Header(subject, 'utf-8')
# Define todays date
from datetime import datetime, timedelta
now = datetime.now()
# Define what date is in one day
delta = timedelta(days=1)
tomorrow = now + delta
# Convert tomorrow date into string
tomorrow = tomorrow.strftime('%d.%m.%Y')
# Open the excel file with the deadlines on the first sheet
wb = openpyxl.load_workbook('Tasks deadlines.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
colC = ws['C4']
colC = str(colC)
colA = ws['A']
# Read the data in column A
for column in ws.columns:
	for colA in column:
# Compare the data in column A to the tomorrow's date
		if colA.value == tomorrow:
# Enter smtp of gmail, login my email, send email, exit smtp
			smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
			smtpObj.starttls()
			smtpObj.login('natalya.agilent@gmail.com', 'YfnfkmzYfnfkmz321')
			smtpObj.sendmail("natalya.agilent@gmail.com", "%s" % colC, msg.as_string())
			smtpObj.quit()
			print("Email sent")
		else:
			print ("Deadline is not tomorrow Do not send anything")
