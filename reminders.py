# This code checks if in the Excel sheet there are deadlines scheduled for the next day and sends automatic reminders. 
import schedule
import time
import smtplib
import openpyxl
import sys
from datetime import datetime, timedelta
now = datetime.now()
# Define what date is in one day
delta = timedelta(days=1)
tomorrow = now + delta
# Convert tomorrow date into string
tomorrow = tomorrow.strftime('%d.%m.%Y')
# We use email and MIME handling packages to create a nice looking email with subject line
wb = openpyxl.load_workbook('Tasks deadlines.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
# Read data in all rows of column A
for i in range(1, 100):
    cellVal = ws.cell(row=i, column=1).value
# If date in column A corresponds to tomorrow's date, read data in the same row, but in column C and define this data as "rowVal"
    if cellVal == tomorrow:
        recAdress = ws.cell(row=i, column=3).value
# also read data in the same row, but in column B and define this data as "recName"
        recName = ws.cell(row=i, column=2).value
        name = "%s" % recName # Create new variable "name"
 # We use email and MIME handling packages to create a nice looking email with subject line
        from email.mime.text import MIMEText
        from email.header import Header
        subject = u'Срок сдачи по одному из проектов в Cloudwords'
# In the next line we create body of the email by adding "name" variable to the string. It is done like that in order to eliminate the encoding errors
        body = name + u', добрый день!\nЗавтра истекает срок сдачи редактуры по одному из проектов. Прошу сообщить, если по каким-то причинам проект не будет сдан в срок. Все проекты можно посмотреть здесь https://app.cloudwords.com\nЗаранее большое спасибо!\nНаталья Курышева.'  
        msg = MIMEText(body, 'plain' 'utf-8')
        msg['Subject'] = Header(subject, 'utf-8')
# Read data in all rows of column A and send emails if the date is tomorrow's date. Use rowVal as a reciever's address
        smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
        smtpObj.starttls()
        smtpObj.login('natalya.agilent@gmail.com', #Password here in '')
        smtpObj.sendmail("natalya.agilent@gmail.com", "%s" % recAdress, msg.as_string()) # in this line on the second place use rowVal as a variable
        smtpObj.quit()
        print("Email sent.")
